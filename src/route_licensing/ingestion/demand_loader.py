"""
demand_loader.py
================
Loads, normalises, and caches passenger demand data from an OD demand
Excel file uploaded by the licensing officer or analyst.

The demand file maps (origin stop, destination stop, day type, time band)
→ estimated passengers per hour.  Missing pairs are handled gracefully —
the analysis continues normally and the UI shows "No data" for those pairs.

Disk cache
----------
Parsed demand is written to data/demand/demand_cache.json so the Excel
file is not re-parsed on every server restart.  The cache is regenerated
whenever a new demand file is uploaded.

Day types
---------
weekday   — Monday to Friday
saturday  — Saturday
sunday    — Sunday / public holiday

Time bands  (must match decision_engine._get_service_band)
----------
am_peak   — before 09:30
midday    — 09:30 – 16:00
pm_peak   — 16:00 – 19:00
evening   — after 19:00
"""

from __future__ import annotations

import json
import logging
import os
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)

CACHE_PATH   = Path("data/demand/demand_cache.json")
VALID_DAY_TYPES  = {"weekday", "saturday", "sunday"}
VALID_TIME_BANDS = {"am_peak", "midday", "pm_peak", "evening"}

# Column aliases accepted in the uploaded Excel (case-insensitive).
_COL_ALIASES: dict[str, list[str]] = {
    "stop_id_from":   ["stop_id_from", "from_stop", "from_stop_id", "origin"],
    "stop_id_to":     ["stop_id_to",   "to_stop",   "to_stop_id",   "destination"],
    "day_type":       ["day_type",     "day",        "service_day"],
    "time_band":      ["time_band",    "band",       "period"],
    "pax_per_hour":   ["pax_per_hour", "passengers", "demand", "pax_ph", "boardings"],
}

# --------------------------------------------------------------------------
# Public API
# --------------------------------------------------------------------------

DemandIndex = dict[tuple[str, str, str, str], float]
"""(from_stop_id, to_stop_id, day_type, time_band) → passengers per hour"""


def load_demand_from_excel(
    file_path: str,
    stop_code_map: dict | None = None,
    all_stop_ids: frozenset | None = None,
) -> DemandIndex:
    """
    Parses the demand Excel file, normalises stop IDs, and returns a
    DemandIndex dict.  Also writes the result to the disk cache.

    Accepts the first sheet named 'OD Demand' (case-insensitive), or the
    first sheet if no such name exists.

    Parameters
    ----------
    file_path       : Path to the .xlsx demand file.
    stop_code_map   : Optional NaPTAN stop_code → full stop_id lookup.
    all_stop_ids    : Optional set of valid GTFS stop IDs for validation.
    """
    df = _read_sheet(file_path)
    df = _normalise_columns(df)
    df = _normalise_stop_ids(df, stop_code_map, all_stop_ids)
    df = _validate_values(df)

    index: DemandIndex = {}
    for row in df.itertuples(index=False):
        key = (row.stop_id_from, row.stop_id_to, row.day_type, row.time_band)
        index[key] = float(row.pax_per_hour)

    logger.info(
        "Demand data loaded: %d valid records from '%s'.",
        len(index), file_path,
    )
    _write_cache(index, source_file=os.path.basename(file_path))
    return index


def load_demand_from_cache() -> DemandIndex | None:
    """
    Loads the demand index from the JSON disk cache if it exists.
    Returns None if no cache is present.
    """
    if not CACHE_PATH.exists():
        return None
    try:
        with CACHE_PATH.open(encoding="utf-8") as f:
            raw = json.load(f)
        index: DemandIndex = {
            (r[0], r[1], r[2], r[3]): float(r[4])
            for r in raw.get("records", [])
        }
        logger.info(
            "Demand cache loaded: %d records (source: %s, cached: %s).",
            len(index),
            raw.get("meta", {}).get("source_file", "?"),
            raw.get("meta", {}).get("loaded_at", "?")[:10],
        )
        return index
    except Exception as exc:
        logger.warning("Could not load demand cache: %s.", exc)
        return None


def lookup_demand(
    index: DemandIndex,
    from_id: str,
    to_id: str,
    day_type: str,
    time_band: str,
) -> float | None:
    """
    Returns the demand (passengers per hour) for the given OD pair and
    period, or None if no data is available.

    Fallback chain:
    1. Exact match  (from, to, day_type, time_band)
    2. Day-type wildcard — any record for (from, to, *, time_band)
    3. None
    """
    exact = index.get((from_id, to_id, day_type, time_band))
    if exact is not None:
        return exact
    # Fallback: any day_type for same pair + band
    for dt in VALID_DAY_TYPES - {day_type}:
        val = index.get((from_id, to_id, dt, time_band))
        if val is not None:
            return val
    return None


def infer_day_type(section_day_groups: str) -> str:
    """
    Infers weekday / saturday / sunday from the section_day_groups string
    stored by the parser (a JSON-ish list of day strings).
    """
    try:
        text = section_day_groups.replace("'", '"').lower()
        groups = json.loads(text)
        combined = " ".join(groups).lower()
    except Exception:
        combined = section_day_groups.lower()

    if "sat" in combined:
        return "saturday"
    if "sun" in combined or "bank" in combined or "public" in combined:
        return "sunday"
    return "weekday"


def get_cache_meta() -> dict:
    """Returns metadata about the current demand cache, or empty dict."""
    if not CACHE_PATH.exists():
        return {}
    try:
        with CACHE_PATH.open(encoding="utf-8") as f:
            raw = json.load(f)
        return raw.get("meta", {})
    except Exception:
        return {}


# --------------------------------------------------------------------------
# Template generation
# --------------------------------------------------------------------------

TEMPLATE_SAMPLE_DATA = [
    # (from_stop_id, to_stop_id, day_type, time_band, pax_per_hour, source, notes)
    ("8310B1018001", "8220B1350301", "weekday", "am_peak",  245, "AFC/Ticketing", "Main corridor, high demand"),
    ("8310B1018001", "8220B1350301", "weekday", "midday",    98, "AFC/Ticketing", ""),
    ("8310B1018001", "8220B1350301", "weekday", "pm_peak",  187, "AFC/Ticketing", "Return commuter peak"),
    ("8310B1018001", "8220B1350301", "weekday", "evening",   34, "Estimated",     ""),
    ("8310B1018001", "8220B1350301", "saturday", "am_peak",  82, "Manual Count",  "Weekend demand lower"),
    ("8310B1018001", "8220B1350301", "saturday", "midday",  110, "Manual Count",  "Leisure/shopping trips"),
    ("8310B1018001", "8220B1350301", "sunday",  "midday",    67, "Estimated",     ""),
    ("8310B1013601", "8220B1350301", "weekday", "am_peak",  156, "AFC/Ticketing", ""),
    ("8310B1013601", "8220B1350301", "weekday", "pm_peak",  143, "AFC/Ticketing", ""),
    ("8310B1013601", "8220B1350301", "saturday", "midday",   72, "Manual Count",  ""),
    ("8240B101281",  "8220B1350301", "weekday", "am_peak",   89, "Model",         "Modelled, validate with count"),
    ("8240B101281",  "8220B1350301", "weekday", "pm_peak",   78, "Model",         ""),
    ("8220B1350301", "8310B1018001", "weekday", "am_peak",   45, "AFC/Ticketing", "Return leg, lower inbound AM"),
    ("8220B1350301", "8310B1018001", "weekday", "pm_peak",  198, "AFC/Ticketing", "Heavy return commuter flow"),
    ("8220B1350301", "8310B1013601", "weekday", "pm_peak",  112, "AFC/Ticketing", ""),
    # Deliberately leave some OD pairs missing to demonstrate partial data
]


def generate_demand_template(output_path: str) -> None:
    """
    Writes a formatted Excel demand template to output_path.
    Includes an Instructions sheet, the OD Demand data sheet, and a
    Valid Values reference sheet.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    # ── Sheet 1: Instructions ──────────────────────────────────────────────
    ws_info = wb.active
    ws_info.title = "Instructions"
    _style_instructions(ws_info)

    # ── Sheet 2: OD Demand (data entry) ───────────────────────────────────
    ws_data = wb.create_sheet("OD Demand")
    _style_data_sheet(ws_data)

    # ── Sheet 3: Valid Values ──────────────────────────────────────────────
    ws_ref = wb.create_sheet("Valid Values")
    _style_reference_sheet(ws_ref)

    wb.save(output_path)
    logger.info("Demand template written to '%s'.", output_path)


# --------------------------------------------------------------------------
# Internal helpers
# --------------------------------------------------------------------------

def _read_sheet(file_path: str) -> pd.DataFrame:
    xl = pd.ExcelFile(file_path)
    sheet_name = next(
        (s for s in xl.sheet_names if "od" in s.lower() or "demand" in s.lower()),
        xl.sheet_names[0],
    )
    df = xl.parse(sheet_name, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    return df.dropna(how="all")


def _normalise_columns(df: pd.DataFrame) -> pd.DataFrame:
    col_map: dict[str, str] = {}
    lower_cols = {c.lower().replace(" ", "_"): c for c in df.columns}
    for canonical, aliases in _COL_ALIASES.items():
        for alias in aliases:
            if alias in lower_cols:
                col_map[lower_cols[alias]] = canonical
                break
    missing = [c for c in _COL_ALIASES if c not in col_map.values()
               and c != "pax_per_hour"]
    if "pax_per_hour" not in col_map.values():
        missing.append("pax_per_hour")
    if missing:
        raise ValueError(
            f"Demand file is missing required columns: {missing}. "
            f"Found columns: {list(df.columns)}"
        )
    return df.rename(columns=col_map)[list(_COL_ALIASES.keys())]


def _normalise_stop_ids(
    df: pd.DataFrame,
    stop_code_map: dict | None,
    all_stop_ids: frozenset | None,
) -> pd.DataFrame:
    if not stop_code_map and not all_stop_ids:
        return df
    known   = all_stop_ids or frozenset()
    sc_map  = stop_code_map or {}

    def _resolve(sid: str) -> str:
        if not isinstance(sid, str):
            return str(sid)
        sid = sid.strip()
        if sid in known:
            return sid
        return sc_map.get(sid, sid)

    df = df.copy()
    df["stop_id_from"] = df["stop_id_from"].map(_resolve)
    df["stop_id_to"]   = df["stop_id_to"].map(_resolve)
    return df


def _validate_values(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["day_type"]  = df["day_type"].str.strip().str.lower()
    df["time_band"] = df["time_band"].str.strip().str.lower()

    bad_day  = ~df["day_type"].isin(VALID_DAY_TYPES)
    bad_band = ~df["time_band"].isin(VALID_TIME_BANDS)
    if bad_day.any():
        logger.warning(
            "Dropping %d rows with invalid day_type values: %s",
            bad_day.sum(), df.loc[bad_day, "day_type"].unique().tolist()
        )
    if bad_band.any():
        logger.warning(
            "Dropping %d rows with invalid time_band values: %s",
            bad_band.sum(), df.loc[bad_band, "time_band"].unique().tolist()
        )

    df = df[~bad_day & ~bad_band].copy()
    df["pax_per_hour"] = pd.to_numeric(df["pax_per_hour"], errors="coerce")
    dropped = df["pax_per_hour"].isna().sum()
    if dropped:
        logger.warning("Dropping %d rows with non-numeric pax_per_hour.", dropped)
    return df.dropna(subset=["pax_per_hour", "stop_id_from", "stop_id_to"])


def _write_cache(index: DemandIndex, source_file: str) -> None:
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "meta": {
            "source_file": source_file,
            "loaded_at":   datetime.now(timezone.utc).isoformat(),
            "record_count": len(index),
        },
        "records": [
            [k[0], k[1], k[2], k[3], v]
            for k, v in index.items()
        ],
    }
    with CACHE_PATH.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    logger.info("Demand cache written: %d records → %s", len(index), CACHE_PATH)


# --------------------------------------------------------------------------
# Excel template styling helpers
# --------------------------------------------------------------------------

def _style_instructions(ws) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    TEAL  = "FF00897B"
    LIGHT = "FFF1F8F7"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70

    rows = [
        ("Passenger Demand Template", "", True, TEAL, "FFFFFFFF"),
        ("", "", False, None, None),
        ("Purpose", "Provide estimated passenger demand for OD pairs on the proposed route.", False, None, None),
        ("", "This data enriches the route licensing analysis with demand context.", False, None, None),
        ("", "", False, None, None),
        ("Sheet to fill in", "Go to the 'OD Demand' sheet and add your data.", False, None, None),
        ("", "", False, None, None),
        ("Required columns", "", True, LIGHT, None),
        ("stop_id_from", "GTFS stop ID of the boarding stop (origin). Also accepts stop_code.", False, None, None),
        ("stop_id_to",   "GTFS stop ID of the alighting stop (destination). Also accepts stop_code.", False, None, None),
        ("day_type",     "One of: weekday / saturday / sunday", False, None, None),
        ("time_band",    "One of: am_peak / midday / pm_peak / evening", False, None, None),
        ("pax_per_hour", "Estimated passengers boarding per hour for this OD pair and period.", False, None, None),
        ("", "", False, None, None),
        ("Optional columns", "", True, LIGHT, None),
        ("data_source",  "Origin of data: AFC/Ticketing, Manual Count, Model, Estimated, etc.", False, None, None),
        ("survey_date",  "Date the data was collected (YYYY-MM-DD).", False, None, None),
        ("notes",        "Free text notes about this record.", False, None, None),
        ("", "", False, None, None),
        ("Partial data OK", "You do not need to provide data for every stop pair.", False, None, None),
        ("",               "Missing pairs are shown as 'No data' in the analysis — they do not affect scoring.", False, None, None),
        ("", "", False, None, None),
        ("Stop IDs",    "Use the GTFS stop IDs (e.g. 8310B1018001) or the short stop_code (e.g. 158131).", False, None, None),
        ("",            "The system will normalise stop_code values automatically.", False, None, None),
    ]

    for i, (label, value, bold, fill_hex, font_hex) in enumerate(rows, start=1):
        cell_a = ws.cell(row=i, column=1, value=label)
        cell_b = ws.cell(row=i, column=2, value=value)
        for cell in (cell_a, cell_b):
            cell.font = Font(bold=bold, color=font_hex or "FF000000", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill_hex:
                cell.fill = PatternFill("solid", fgColor=fill_hex)

    ws.row_dimensions[1].height = 22


def _style_data_sheet(ws) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="FF00897B")
    ALT_FILL    = PatternFill("solid", fgColor="FFF1F8F7")
    HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=10)
    BODY_FONT   = Font(size=10)
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left", vertical="center")
    thin        = Side(style="thin", color="FFB2DFDB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "stop_id_from", "stop_id_to", "day_type", "time_band",
        "pax_per_hour", "data_source", "survey_date", "notes",
    ]
    col_widths = [18, 18, 14, 14, 16, 18, 16, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font    = HEADER_FONT
        cell.fill    = HEADER_FILL
        cell.alignment = CENTER
        cell.border  = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20

    for row_idx, record in enumerate(TEMPLATE_SAMPLE_DATA, start=2):
        from_id, to_id, day_type, time_band, pax, source, notes = record
        values = [from_id, to_id, day_type, time_band, pax, source, "", notes]
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = BODY_FONT
            cell.border    = border
            cell.alignment = CENTER if col in (3, 4, 5) else LEFT
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{1 + len(TEMPLATE_SAMPLE_DATA)}"


def _style_reference_sheet(ws) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    TEAL  = "FF00897B"
    LIGHT = "FFF1F8F7"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 50

    sections = [
        ("day_type values", [
            ("weekday",  "Monday to Friday"),
            ("saturday", "Saturday"),
            ("sunday",   "Sunday and public holidays"),
        ]),
        ("time_band values", [
            ("am_peak", "Before 09:30"),
            ("midday",  "09:30 to 16:00"),
            ("pm_peak", "16:00 to 19:00"),
            ("evening", "After 19:00"),
        ]),
        ("data_source suggestions", [
            ("AFC/Ticketing",  "Automated Fare Collection or ticketing system data"),
            ("Manual Count",   "Passenger count survey"),
            ("Model",          "Demand model output"),
            ("Estimated",      "Analyst estimate"),
        ]),
    ]

    row = 1
    for title, items in sections:
        h = ws.cell(row=row, column=1, value=title)
        h.font = Font(bold=True, color="FFFFFFFF", size=10)
        h.fill = PatternFill("solid", fgColor=TEAL)
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=TEAL)
        row += 1
        for val, desc in items:
            c1 = ws.cell(row=row, column=1, value=val)
            c2 = ws.cell(row=row, column=2, value=desc)
            c1.font = Font(bold=True, size=10)
            c2.font = Font(size=10)
            if row % 2 == 0:
                for c in (c1, c2):
                    c.fill = PatternFill("solid", fgColor=LIGHT)
            row += 1
        row += 1
