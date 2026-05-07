"""Generates the timetable submission Excel template."""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import time

APP_TEAL  = "009B9E"
APP_NAVY  = "0D2137"
APP_LIME  = "FFE600"
WHITE     = "FFFFFF"
LIGHT_GRAY = "F0F4F8"
LIGHT_TEAL = "E8F7F7"

def hdr_font(bold=True, color=WHITE, size=9):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def cell_font(bold=False, color=APP_NAVY, size=9):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def ctr():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def lft():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Instructions sheet ──────────────────────────────────────────────
wb = openpyxl.Workbook()
ws_info = wb.active
ws_info.title = "Instructions"
ws_info.column_dimensions["A"].width = 30
ws_info.column_dimensions["B"].width = 68

ws_info.merge_cells("A1:B1")
c = ws_info["A1"]
c.value = "Route Licensing — Timetable Submission Template"
c.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
c.fill = fill(APP_NAVY)
c.alignment = ctr()
ws_info.row_dimensions[1].height = 32

ws_info.merge_cells("A2:B2")
c = ws_info["A2"]
c.value = "Read all instructions before filling in the Timetable sheets. Replace placeholder Stop IDs (8380BXXXXXX) with real GTFS stop codes from the loaded feed."
c.font = Font(name="Calibri", italic=True, color=APP_NAVY, size=10)
c.fill = fill(APP_LIME)
c.alignment = ctr()
ws_info.row_dimensions[2].height = 22

SECTION = {"SHEET LAYOUT", "ROW STRUCTURE", "COLUMN DEFINITIONS", "DAY GROUP HEADERS", "IMPORTANT RULES"}

rows = [
    ("", ""),
    ("SHEET LAYOUT", ""),
    ("Timetable_Route_1", 'One sheet per proposed route direction. Rename to describe the direction (e.g. "Kinsale-Cork-Outbound").'),
    ("Timetable_Route_2", "Add more sheets by copying Timetable_Route_1. Each sheet = one direction of travel."),
    ("", ""),
    ("ROW STRUCTURE", ""),
    ('Row 1 — Section Title', 'Single text cell. Free-text name of this direction (e.g. "Kinsale to Cork City"). Must be the ONLY value in the row.'),
    ("Row 2 — Header", 'Fixed: Stop Name | Stop Location | Stop ID | then one or more day-group labels (e.g. "Monday \u2013 Sunday") spanning the departure columns via merged cells.'),
    ("Row 3+ — Stop Rows", "One row per stop. Columns A\u2013C fixed. Columns D onward are departure times as Excel time values (HH:MM)."),
    ("Blank Row", "A blank row ends the section. Leave one blank row between sections if you place multiple directions on one sheet."),
    ("", ""),
    ("COLUMN DEFINITIONS", ""),
    ("Col A \u2014 Stop Name", "Full name of the stop as shown on the stop sign."),
    ("Col B \u2014 Stop Location", "Street / locality (informational only, not used in analysis)."),
    ("Col C \u2014 Stop ID", "GTFS stop code (e.g. 8380B247191). Must match a stop in the current GTFS feed exactly. Invalid IDs are rejected on upload."),
    ("Col D onward \u2014 Times", "Each column is one departure (one trip). Enter as Excel time format HH:MM. Do NOT type as plain text."),
    ("", ""),
    ("DAY GROUP HEADERS", ""),
    ("Merged header cells", 'In Row 2, type the day label (e.g. "Monday \u2013 Friday") and merge it across all columns that belong to that group. The system counts the span automatically.'),
    ("", ""),
    ("IMPORTANT RULES", ""),
    ("Stop IDs must be valid", "Every Stop ID (Col C) must exist in the loaded GTFS static feed. The upload will be rejected if any ID is not found."),
    ("Times must be Excel time cells", 'Do not type "07:30" as plain text. Use Excel time format so the cell stores a real time value. Format cells as HH:MM.'),
    ("At least 2 stops per section", "A section with fewer than 2 unique stops will be rejected."),
    ("Operator name", "The operating company name is entered in the upload form on screen — not in this file."),
    ("Multiple routes", "Submit each proposed route as a separate upload, OR include multiple direction sheets in one file. Both are supported."),
]

for i, (label, desc) in enumerate(rows, start=3):
    ws_info.row_dimensions[i].height = 18
    a = ws_info.cell(row=i, column=1, value=label)
    b = ws_info.cell(row=i, column=2, value=desc)
    if label in SECTION:
        for cell in (a, b):
            cell.font = hdr_font()
            cell.fill = fill(APP_TEAL)
            cell.alignment = lft()
    else:
        a.font = Font(name="Calibri", bold=True, color=APP_NAVY, size=9)
        b.font = Font(name="Calibri", color=APP_NAVY, size=9)
        a.fill = fill(LIGHT_GRAY)
        b.fill = fill(WHITE)
        a.alignment = lft()
        b.alignment = lft()
    a.border = thin()
    b.border = thin()


# ── Timetable sheet builder ─────────────────────────────────────────
def make_sheet(wb, sheet_name, title, stops_data, trips, day_label="Monday \u2013 Sunday"):
    ws = wb.create_sheet(title=sheet_name)

    n_trips = len(trips)
    last_col = 3 + n_trips

    # Row 1 — section title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    c.fill = fill(APP_NAVY)
    c.alignment = ctr()
    ws.row_dimensions[1].height = 26

    # Row 2 — fixed headers
    for col, hdr in enumerate(["Stop Name", "Stop Location", "Stop ID"], start=1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.font = hdr_font()
        c.fill = fill(APP_TEAL)
        c.alignment = ctr()
        c.border = thin()

    # Row 2 — day group merged header
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=last_col)
    c = ws.cell(row=2, column=4, value=day_label)
    c.font = Font(name="Calibri", bold=True, color=APP_NAVY, size=9)
    c.fill = fill(APP_LIME)
    c.alignment = ctr()
    for col in range(4, last_col + 1):
        ws.cell(row=2, column=col).border = thin()
    ws.row_dimensions[2].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    for i in range(n_trips):
        ws.column_dimensions[get_column_letter(4 + i)].width = 9

    # Data rows
    for r, (stop_name, stop_loc, stop_id) in enumerate(stops_data, start=3):
        row_bg = LIGHT_TEAL if r % 2 == 0 else WHITE
        for col, val in enumerate([stop_name, stop_loc, stop_id], start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = cell_font(bold=(col == 1))
            c.fill = fill(row_bg)
            c.alignment = lft()
            c.border = thin()
        for t_idx, t_val in enumerate(trips):
            c = ws.cell(row=r, column=4 + t_idx, value=t_val)
            c.number_format = "HH:MM"
            c.font = Font(name="Calibri", color=APP_TEAL, bold=True, size=9)
            c.fill = fill(row_bg)
            c.alignment = ctr()
            c.border = thin()
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = "D3"
    return ws


# ── Route 1 — Outbound ──────────────────────────────────────────────
outbound_stops = [
    ("Kinsale Town Hall",       "Kinsale",    "8380B247191"),
    ("Kinsale \u2013 Pier Road","Kinsale",    "8380BXXXXXX"),
    ("Belgooly Village",        "Belgooly",   "8380BYYYYYY"),
    ("Halfway",                 "Halfway",    "8380BZZZZZZ"),
    ("Cork \u2013 Parnell Place","Cork City", "8380BAAAAAA"),
]
outbound_trips = [time(7,0), time(9,0), time(11,0), time(13,0), time(15,0), time(17,0), time(19,0)]
make_sheet(wb, "Timetable_Route_1", "Kinsale to Cork City \u2014 Outbound", outbound_stops, outbound_trips)

# ── Route 2 — Inbound ───────────────────────────────────────────────
inbound_stops = [
    ("Cork \u2013 Parnell Place","Cork City", "8380BAAAAAA"),
    ("Halfway",                 "Halfway",    "8380BZZZZZZ"),
    ("Belgooly Village",        "Belgooly",   "8380BYYYYYY"),
    ("Kinsale \u2013 Pier Road","Kinsale",    "8380BXXXXXX"),
    ("Kinsale Town Hall",       "Kinsale",    "8380B247191"),
]
inbound_trips = [time(8,0), time(10,0), time(12,0), time(14,0), time(16,0), time(18,0), time(20,0)]
make_sheet(wb, "Timetable_Route_2", "Cork City to Kinsale \u2014 Inbound", inbound_stops, inbound_trips)

out = os.path.join(os.path.dirname(__file__), "..", "src", "route_licensing", "api", "static", "timetable_template.xlsx")
wb.save(out)
print("Saved:", os.path.abspath(out))
